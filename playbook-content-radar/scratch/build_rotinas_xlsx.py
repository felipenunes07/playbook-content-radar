"""Monta o Excel do cruzamento "5 ROTINAS" (Tally xX8rVJ) x base de conteudo/leads.

Entrada:  scratch/rotinas-dataset.json  (gerado por scratch/rotinas-dataset.ts)
Saida:    scratch/Rotinas - cruzamento Tally x comentaristas.xlsx

Fonte de todos os dados: Supabase xcihctupmfawtawbzwvm (leitura apenas).
Os numeros do Resumo e do ICP sao FORMULAS sobre a aba Inscritos, nao valores fixos.
"""

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

BASE = Path(__file__).resolve().parent
DATA = json.loads((BASE / "rotinas-dataset.json").read_text(encoding="utf-8"))
OUT = BASE / "Rotinas - cruzamento Tally x comentaristas.xlsx"

FONT = "Arial"
INK = "1F2A28"
HEAD_FILL = PatternFill("solid", fgColor="0F6E62")
SUB_FILL = PatternFill("solid", fgColor="E2EEEB")
OK_FILL = PatternFill("solid", fgColor="E2EEEB")
REV_FILL = PatternFill("solid", fgColor="FBECD9")
COLD_FILL = PatternFill("solid", fgColor="EFEFEF")
NOTE_FILL = PatternFill("solid", fgColor="FFF9DB")
THIN = Side(style="thin", color="D6DCD8")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = Workbook()


def style_header(ws, row=1, ncols=None):
    ncols = ncols or ws.max_column
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[row].height = 30


def body_font(ws, first_row=2):
    for row in ws.iter_rows(min_row=first_row, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = Font(name=FONT, size=9, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.border = BOX


def widths(ws, mapping):
    for letter, width in mapping.items():
        ws.column_dimensions[letter].width = width


def write_table(ws, columns, rows, name):
    ws.append([label for label, _ in columns])
    for item in rows:
        ws.append([item.get(key, "") for _, key in columns])
    style_header(ws)
    body_font(ws)
    if ws.max_row > 1:
        ref = f"A1:{get_column_letter(len(columns))}{ws.max_row}"
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(table)
    ws.freeze_panes = "A2"


# ---------------------------------------------------------------- Inscritos

def classificacao(row):
    if row["lixo"]:
        return "Lixo"
    if row["teste"]:
        return "Teste interno"
    return "Real"


INSCRITOS_COLS = [
    ("Classificacao", "classificacao"),
    ("Ja era audiencia?", "audiencia"),
    ("Nome (Tally)", "nome_tally"),
    ("Primeiro nome", "primeiro_nome"),
    ("Sobrenome", "sobrenome"),
    ("E-mail", "email"),
    ("Corporativo", "corporativo"),
    ("Dominio", "dominio"),
    ("Telefone", "telefone"),
    ("Inscrito em", "inscrito_em"),
    ("Inscricoes neste form", "inscricoes_neste_form"),
    ("Match", "match_status"),
    ("Metodo do match", "match_metodo"),
    ("Confianca", "match_confianca"),
    ("Evidencia", "match_evidencia"),
    ("Leads candidatos", "leads_candidatos"),
    ("LinkedIn", "linkedin"),
    ("Nome no LinkedIn", "nome_linkedin"),
    ("Cargo / headline", "cargo_ou_headline"),
    ("Senioridade", "senioridade"),
    ("Area", "area"),
    ("Empresa", "empresa"),
    ("Funcionarios", "tamanho_empresa"),
    ("Industria", "industria"),
    ("Cidade", "cidade"),
    ("Estado", "estado"),
    ("Pais", "pais"),
    ("ICP hoje", "icp_status"),
    ("Score ICP", "icp_score"),
    ("Motivo do ICP", "icp_motivo"),
    ("Comentarios", "comentarios"),
    ("1o comentario", "primeiro_comentario"),
    ("Ultimo comentario", "ultimo_comentario"),
    ("Posts comentados", "posts_comentados"),
    ("CTAs que respondeu", "ctas_comentados"),
    ("Outros lead magnets", "outros_lead_magnets"),
    ("Quais lead magnets", "quais_lead_magnets"),
    ("Primeiro contato", "primeiro_contato"),
    ("Dias na base", "dias_na_base"),
    ("Prospectado", "prospectado"),
    ("Canal", "canal_outreach"),
    ("Data do outreach", "data_outreach"),
    ("Agendou reuniao", "agendou_reuniao"),
    ("Telefone ja registrado", "telefone_ja_registrado"),
]

COL = {key: get_column_letter(i + 1) for i, (_, key) in enumerate(INSCRITOS_COLS)}

ORDEM_AUDIENCIA = {"Comentou em post": 0, "Baixou outro lead magnet": 1, "Novo": 2}
ORDEM_MATCH = {"MATCHED": 0, "MATCHED_NO_PHONE": 1, "REVIEW": 2, "NOT_FOUND": 3}

inscritos = []
for row in DATA["inscritos"]:
    item = dict(row)
    item["classificacao"] = classificacao(row)
    item["cargo_ou_headline"] = row["cargo"] or row["headline"]
    inscritos.append(item)

inscritos.sort(key=lambda r: (
    0 if r["classificacao"] == "Real" else 1,
    ORDEM_AUDIENCIA.get(r["audiencia"], 9),
    ORDEM_MATCH.get(r["match_status"], 9),
    -(r["icp_score"] if isinstance(r["icp_score"], (int, float)) else -1),
    r["nome_tally"].lower(),
))

ws = wb.active
ws.title = "Inscritos"
write_table(ws, INSCRITOS_COLS, inscritos, "Inscritos")
widths(ws, {
    "A": 13, "B": 22, "C": 26, "D": 14, "E": 16, "F": 34, "G": 11, "H": 24, "I": 16,
    "J": 17, "K": 9, "L": 17, "M": 24, "N": 10, "O": 26, "P": 9, "Q": 42, "R": 24,
    "S": 40, "T": 13, "U": 13, "V": 26, "W": 12, "X": 26, "Y": 16, "Z": 14, "AA": 12,
    "AB": 13, "AC": 10, "AD": 46, "AE": 12, "AF": 14, "AG": 16, "AH": 34, "AI": 18,
    "AJ": 9, "AK": 44, "AL": 15, "AM": 12, "AN": 13, "AO": 12, "AP": 15, "AQ": 22, "AR": 22,
})
ws.freeze_panes = "D2"

# realces por status de match e por classificacao
for i, item in enumerate(inscritos, start=2):
    fill = None
    if item["classificacao"] != "Real":
        fill = COLD_FILL
    elif item["match_status"] in ("MATCHED", "MATCHED_NO_PHONE"):
        fill = OK_FILL
    elif item["match_status"] == "REVIEW":
        fill = REV_FILL
    if fill:
        for col in ("A", "B", "L", "M", "N"):
            ws[f"{col}{i}"].fill = fill
    ws[f"{COL['match_confianca']}{i}"].number_format = "0.00"
    for key in ("tamanho_empresa", "icp_score", "comentarios", "outros_lead_magnets", "dias_na_base",
                "inscricoes_neste_form", "leads_candidatos"):
        ws[f"{COL[key]}{i}"].number_format = "0"

LAST = ws.max_row


def contar(*pares):
    """COUNTIFS sobre Inscritos, sempre restrito a Classificacao = Real."""
    partes = [f"Inscritos!${COL['classificacao']}$2:${COL['classificacao']}${LAST},\"Real\""]
    for key, criterio in pares:
        partes.append(f"Inscritos!${COL[key]}$2:${COL[key]}${LAST},{criterio}")
    return "=COUNTIFS(" + ",".join(partes) + ")"


# ---------------------------------------------------------------- Resumo

ws = wb.create_sheet("Resumo")
ws["A1"] = "Cruzamento: Tally \"5 ROTINAS PARA AUTOMATIZAR SUA OPERACAO\" x base de conteudo"
ws["A1"].font = Font(name=FONT, size=15, bold=True, color=INK)
ws["A2"] = f"Formulario {DATA['formId']} | dados lidos do Supabase xcihctupmfawtawbzwvm em {DATA['geradoEm']} | leitura apenas"
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5C6B66")

linhas = [
    ("A ORIGEM", "", ""),
    ("Inscricoes no formulario", f"=COUNTA(Inscritos!${COL['email']}$2:${COL['email']}${LAST})",
     "Uma linha por submissao unica de e-mail; duplicatas do mesmo e-mail foram agrupadas."),
    ("Pessoas reais", contar(), "Exclui 2 linhas de lixo (ASDF, Teste) e 2 inscricoes de teste interno."),
    ("Com telefone", contar(("telefone", "\"+*\"")), "Este formulario coleta telefone - todas as inscricoes vieram com numero."),
    ("Com e-mail corporativo", contar(("corporativo", "\"sim\"")), "Dominio proprio: da a empresa mesmo sem perfil de LinkedIn."),
    ("", "", ""),
    ("JA ERAM NOSSA AUDIENCIA", "", ""),
    ("Comentaram em post nosso", contar(("audiencia", "\"Comentou em post\"")),
     "Cruzado por nome contra 3.249 comentarios coletados de 254 posts."),
    ("  dos quais confirmados", contar(("match_status", "\"MATCHED\"")),
     "Nome exato + especificidade ou dominio de e-mail batendo com a empresa."),
    ("  dos quais a revisar", contar(("match_status", "\"REVIEW\"")),
     "Nome bate mas falta evidencia independente. Nao afirmar sem conferir."),
    ("Baixaram outro lead magnet", contar(("audiencia", "\"Baixou outro lead magnet\"")),
     "E-mail ja estava na base do Tally, mas nunca comentou."),
    ("Audiencia nova", contar(("audiencia", "\"Novo\"")), "Nem comentario, nem download anterior."),
    ("", "", ""),
    ("O QUE ESSE GRUPO PARECE", "", ""),
    ("Media de funcionarios (casados)",
     f"=IFERROR(AVERAGEIFS(Inscritos!${COL['tamanho_empresa']}$2:${COL['tamanho_empresa']}${LAST},"
     f"Inscritos!${COL['classificacao']}$2:${COL['classificacao']}${LAST},\"Real\","
     f"Inscritos!${COL['tamanho_empresa']}$2:${COL['tamanho_empresa']}${LAST},\">0\"),\"-\")",
     "Puxada pelas tres contas grandes; a mediana esta muito abaixo."),
    ("Qualified pelo score atual", contar(("icp_status", "\"qualified\"")), "Score que o pipeline de prospeccao usa hoje."),
    ("Disqualified pelo score atual", contar(("icp_status", "\"disqualified\"")),
     "Inclui quase todos os mais engajados - ver aba ICP."),
    # Criterio com o "nao" acentuado de proposito: e o valor literal gravado na coluna.
    ("Ja prospectados", contar(("prospectado", "\"<>não\"")), "Cruzado com lead_outreach (136 registros)."),
    ("Agendaram reuniao", contar(("agendou_reuniao", "\"<>não\"")), "Cruzado com lead_magnet_bookings."),
]

row = 4
for label, formula, nota in linhas:
    ws[f"A{row}"] = label
    if not formula and label:
        ws[f"A{row}"].font = Font(name=FONT, size=10, bold=True, color="0F6E62")
        ws[f"A{row}"].fill = SUB_FILL
        ws[f"B{row}"].fill = SUB_FILL
        ws[f"C{row}"].fill = SUB_FILL
    else:
        ws[f"A{row}"].font = Font(name=FONT, size=10, color=INK)
        ws[f"B{row}"] = formula
        ws[f"B{row}"].font = Font(name=FONT, size=11, bold=True, color=INK)
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        ws[f"B{row}"].number_format = "0"
        ws[f"C{row}"] = nota
        ws[f"C{row}"].font = Font(name=FONT, size=9, color="5C6B66")
    row += 1

ws["B18"].number_format = "0"
widths(ws, {"A": 34, "B": 12, "C": 82})
ws["A2"].alignment = Alignment(vertical="center")

aviso_row = row + 1
ws[f"A{aviso_row}"] = "POR QUE O POST DE 25/08 NAO APARECE NA BASE"
ws[f"A{aviso_row}"].font = Font(name=FONT, size=10, bold=True, color="9E5714")
ws[f"A{aviso_row}"].fill = NOTE_FILL
ws[f"B{aviso_row}"].fill = NOTE_FILL
ws[f"C{aviso_row}"].fill = NOTE_FILL
ws[f"A{aviso_row + 1}"] = (
    "O credito da Apify acabou. O coletor de LinkedIn roda 09:30 UTC; em 25/08 rodou e "
    "funcionou, mas o post so subiu ~12:00 - passou depois da coleta. A rodada de 26/08, "
    "que pegaria o post, falhou: \"you will exceed your remaining usage of $0.001342\". "
    "Detalhe completo na aba Diagnostico."
)
ws[f"A{aviso_row + 1}"].font = Font(name=FONT, size=9, color=INK)
ws[f"A{aviso_row + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{aviso_row + 1}:C{aviso_row + 1}")
ws.row_dimensions[aviso_row + 1].height = 46

# ---------------------------------------------------------------- Matches

MATCH_COLS = [
    ("Status", "status"),
    ("Metodo", "metodo"),
    ("Confianca", "confianca"),
    ("Evidencia", "evidencia"),
    ("Candidato", "candidato"),
    ("de", "candidatos_total"),
    ("Tipo de nome", "tipo_de_nome"),
    ("Nome (Tally)", "nome_tally"),
    ("E-mail", "email"),
    ("Telefone", "telefone"),
    ("Inscrito em", "inscrito_em"),
    ("Nome no LinkedIn", "nome_linkedin"),
    ("LinkedIn", "linkedin"),
    ("Cargo", "cargo"),
    ("Empresa", "empresa"),
    ("Funcionarios", "tamanho_empresa"),
    ("ICP", "icp_status"),
    ("Score", "icp_score"),
    ("Comentarios", "comentarios"),
    ("Posts comentados", "posts_comentados"),
]

matches = sorted(
    DATA["matches"],
    key=lambda r: (ORDEM_MATCH.get(r["status"], 9), -r["confianca"], r["email"], r["candidato"]),
)
ws = wb.create_sheet("Matches")
write_table(ws, MATCH_COLS, matches, "Matches")
widths(ws, {
    "A": 18, "B": 28, "C": 10, "D": 26, "E": 10, "F": 5, "G": 16, "H": 24, "I": 32,
    "J": 16, "K": 17, "L": 26, "M": 46, "N": 40, "O": 26, "P": 12, "Q": 13, "R": 8,
    "S": 12, "T": 34,
})
for i, item in enumerate(matches, start=2):
    fill = OK_FILL if item["status"].startswith("MATCHED") else REV_FILL
    for col in ("A", "B", "C"):
        ws[f"{col}{i}"].fill = fill
    ws[f"C{i}"].number_format = "0.00"

nota = ws.max_row + 2
ws[f"A{nota}"] = (
    "Uma linha por par inscricao x lead candidato. Quando 'de' e maior que 1, a mesma inscricao "
    "do Tally casa com mais de um lead - e o caso do \"Rafael Silva\", com tres pessoas reais "
    "disputando a mesma inscricao. Regra do matcher: nome sozinho so decide com 3+ tokens e "
    "candidato unico; abaixo disso exige e-mail corporativo batendo com a empresa ou inscricao "
    "no formulario do post comentado."
)
ws[f"A{nota}"].font = Font(name=FONT, size=9, italic=True, color="5C6B66")
ws[f"A{nota}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{nota}:H{nota}")
ws.row_dimensions[nota].height = 60

# ---------------------------------------------------------------- Comentarios

COMENT_COLS = [
    ("Nome no LinkedIn", "nome_linkedin"),
    ("Match", "match_status"),
    ("Cargo", "cargo"),
    ("Empresa", "empresa"),
    ("Comentou em", "comentado_em"),
    ("Post de", "post_publicado_em"),
    ("Autor do post", "post_autor"),
    ("CTA do post", "post_cta"),
    ("Gancho do post", "post_hook"),
    ("O que comentou", "comentario"),
]
comentarios = sorted(DATA["comentarios"], key=lambda r: (r["nome_linkedin"].lower(), r["comentado_em"]))
ws = wb.create_sheet("Comentarios")
write_table(ws, COMENT_COLS, comentarios, "Comentarios")
widths(ws, {"A": 26, "B": 18, "C": 38, "D": 24, "E": 13, "F": 12, "G": 16, "H": 12, "I": 52, "J": 80})
for i in range(2, ws.max_row + 1):
    ws[f"J{i}"].alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------- Historico

HIST_COLS = [
    ("Nome (Tally)", "nome_tally"),
    ("E-mail", "email"),
    ("Ja era audiencia?", "audiencia"),
    ("Lead magnet baixado antes", "lead_magnet"),
    ("Form ID", "form_id"),
    ("Baixado em", "baixado_em"),
    ("Telefone naquele form", "telefone"),
]
historico = sorted(DATA["historico"], key=lambda r: (r["nome_tally"].lower(), r["baixado_em"]))
ws = wb.create_sheet("Historico Tally")
write_table(ws, HIST_COLS, historico, "HistoricoTally")
widths(ws, {"A": 26, "B": 34, "C": 24, "D": 44, "E": 10, "F": 17, "G": 16})

# ---------------------------------------------------------------- ICP

ws = wb.create_sheet("ICP")
ws["A1"] = "Perfil de quem se inscreveu E ja tinha comentado em post nosso"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
ws["A2"] = "Contagens calculadas por formula sobre a aba Inscritos, sempre restritas a Classificacao = Real."
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5C6B66")

ws.append([])
ws.append(["Corte", "Pessoas", "Leitura"])
style_header(ws, row=4, ncols=3)

COMENTOU = ("audiencia", "\"Comentou em post\"")

cortes = [
    ("Area classificada como vendas", contar(COMENTOU, ("area", "\"vendas\"")),
     "Classificacao automatica; o titulo real mostra mais - ver coluna Cargo / headline."),
    ("Senioridade c-level", contar(COMENTOU, ("senioridade", "\"c-level\"")), "Fundadores de operacao enxuta."),
    ("Senioridade diretoria ou gerencia",
     "=" + contar(COMENTOU, ("senioridade", "\"diretoria\""))[1:] + "+" + contar(COMENTOU, ("senioridade", "\"gerencia\""))[1:],
     "Camada que decide processo comercial."),
    ("Senioridade operacional ou coordenacao",
     "=" + contar(COMENTOU, ("senioridade", "\"operacional\""))[1:] + "+" + contar(COMENTOU, ("senioridade", "\"coordenacao\""))[1:],
     "Quem executa a rotina que o post prometia automatizar."),
    ("Senioridade nao identificada", contar(COMENTOU, ("senioridade", "\"desconhecido\"")),
     "Perfil sem cargo estruturado no scrape - da para ler pelo headline."),
    ("Empresa com mais de 1.000 funcionarios", contar(COMENTOU, ("tamanho_empresa", "\">1000\"")),
     "HubSpot, Keeta e Cantu - a excecao, nao a regra."),
    ("Empresa entre 11 e 1.000", "=" + contar(COMENTOU, ("tamanho_empresa", "\">=11\""))[1:] + "-" + contar(COMENTOU, ("tamanho_empresa", "\">1000\""))[1:],
     "HostGator (316), Vkron (27), Senior Mega SPO (11). Faiston (201) aparece como candidato "
     "alternativo da colisao \"Rafael Silva\" - conta na aba Matches, nao aqui."),
    ("Empresa com 10 ou menos", "=" + contar(COMENTOU, ("tamanho_empresa", "\">0\""))[1:] + "-" + contar(COMENTOU, ("tamanho_empresa", "\">10\""))[1:],
     "Consultores e fundadores solo."),
    ("Sem empresa identificada", "=" + contar(COMENTOU)[1:] + "-" + contar(COMENTOU, ("tamanho_empresa", "\">0\""))[1:],
     "Perfil sem vinculo atual no scrape."),
    ("Qualified pelo score atual", contar(COMENTOU, ("icp_status", "\"qualified\"")), "Marina Lemos (95) e Rodrigo Nascimento (55)."),
    ("Disqualified pelo score atual", contar(COMENTOU, ("icp_status", "\"disqualified\"")), "Score entre 30 e 45 na maioria."),
    ("Comentou em 2 ou mais posts", contar(COMENTOU, ("comentarios", "\">=2\"")), "Recorrencia, nao acaso."),
    ("Tambem baixou outro lead magnet", contar(COMENTOU, ("outros_lead_magnets", "\">0\"")),
     "Comentou E baixou material antes: o sinal mais forte de intencao que temos."),
]

for label, formula, leitura in cortes:
    ws.append([label, formula, leitura])

for i in range(5, ws.max_row + 1):
    ws.cell(row=i, column=1).font = Font(name=FONT, size=10, color=INK)
    ws.cell(row=i, column=2).font = Font(name=FONT, size=11, bold=True, color=INK)
    ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
    ws.cell(row=i, column=2).number_format = "0"
    ws.cell(row=i, column=3).font = Font(name=FONT, size=9, color="5C6B66")
    for col in range(1, 4):
        ws.cell(row=i, column=col).border = BOX

tensao = ws.max_row + 2
ws[f"A{tensao}"] = "A TENSAO QUE VALE DECIDIR"
ws[f"A{tensao}"].font = Font(name=FONT, size=10, bold=True, color="9E5714")
ws[f"A{tensao}"].fill = NOTE_FILL
ws[f"B{tensao}"].fill = NOTE_FILL
ws[f"C{tensao}"].fill = NOTE_FILL
ws[f"A{tensao + 1}"] = (
    "Quase todo mundo que comentou, voltou e baixou material de novo esta marcado como "
    "disqualified, com score entre 30 e 45. Quem responde a esse tema e operador de vendas em "
    "empresa pequena ou media - SDR, inside sales, sales ops, consultor solo, fundador de time "
    "de 2 a 30 pessoas. E o mesmo publico do post de 11/08 sobre rotina de SDR, que voltou "
    "agora. Se esse perfil e ICP, o score precisa parar de puni-lo. Se nao e, a pauta que mais "
    "engaja e a que menos vende - e isso e decisao de conteudo, nao de dado."
)
ws[f"A{tensao + 1}"].font = Font(name=FONT, size=9, color=INK)
ws[f"A{tensao + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{tensao + 1}:C{tensao + 1}")
ws.row_dimensions[tensao + 1].height = 76
widths(ws, {"A": 40, "B": 11, "C": 74})

# ---------------------------------------------------------------- Enriquecer

ws = wb.create_sheet("Enriquecer")
ws["A1"] = "Inscritos com e-mail corporativo e sem perfil de LinkedIn na base"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
ws["A2"] = "O dominio entrega a empresa. Caminho mais curto para nome novo com contexto."
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5C6B66")
ws.append([])
ws.append(["Nome", "E-mail", "Dominio", "Telefone", "Inscrito em", "Observacao"])
style_header(ws, row=4, ncols=6)

SUSPEITO = {
    "hotmail.con": "Erro de digitacao de hotmail.com - classificado como corporativo por engano.",
    "simplelogin.com": "Servico de e-mail descartavel; nao identifica empresa.",
}

para_enriquecer = [
    r for r in inscritos
    if r["classificacao"] == "Real" and r["corporativo"] == "sim" and not r["linkedin"]
]
para_enriquecer.sort(key=lambda r: r["dominio"])

for item in para_enriquecer:
    obs = SUSPEITO.get(item["dominio"], "")
    if item["telefone"].endswith("999999999") or item["telefone"].endswith("99999999"):
        obs = (obs + " Telefone aparenta ser falso.").strip()
    ws.append([item["nome_tally"], item["email"], item["dominio"], item["telefone"], item["inscrito_em"], obs])

body_font(ws, first_row=5)
widths(ws, {"A": 26, "B": 40, "C": 30, "D": 17, "E": 17, "F": 72})
for i in range(5, ws.max_row + 1):
    ws.cell(row=i, column=6).alignment = Alignment(wrap_text=True, vertical="top")

# ---------------------------------------------------------------- Diagnostico

ws = wb.create_sheet("Diagnostico")
ws["A1"] = "Por que o post de 25/08 nao entrou na base"
ws["A1"].font = Font(name=FONT, size=14, bold=True, color=INK)
ws["A2"] = "Evidencia lida de public.collection_runs e de cron.job no Supabase, em 26/08/2026."
ws["A2"].font = Font(name=FONT, size=9, italic=True, color="5C6B66")

ws.append([])
ws.append(["Quando", "Coletor", "Resultado", "O que aconteceu"])
style_header(ws, row=4, ncols=4)

# Fonte: collection_runs (source apify_linkedin / apify_instagram / apify_youtube) e cron.job.
DIAGNOSTICO = [
    ("25/08 09:30 UTC", "collect-linkedin", "sucesso - 7 itens",
     "Rodou no horario e funcionou. O post de rotinas so subiu por volta de 12:00 UTC (a primeira "
     "inscricao no Tally e 12:04), ou seja, depois desta coleta."),
    ("25/08 14:00 UTC", "catch-up", "nao disparou",
     "O job de recuperacao so roda se NAO houve coleta bem-sucedida nas ultimas 12h. A de 09:30 "
     "tinha dado certo, entao ele corretamente pulou - e o post perdeu a segunda chance do dia."),
    ("26/08 09:00 UTC", "collect-youtube", "parcial - actor ABORTED",
     "Primeiro sinal do problema do dia."),
    ("26/08 09:30 UTC", "collect-linkedin", "FALHOU - 3 contas",
     "\"By launching this job you will exceed your remaining usage of $0.001342. Consider upgrading "
     "to a paid plan\". Credito da Apify esgotado. Era esta rodada que pegaria o post."),
    ("26/08 10:00 UTC", "collect-instagram", "FALHOU", "Mesmo erro de credito da Apify."),
    ("18/08", "todos", "FALHOU - 172 execucoes",
     "\"Monthly usage hard limit exceeded\". Voltou a funcionar em 19/08, o que sugere que o ciclo de "
     "cobranca da Apify reseta dia 19. Se for isso, sem recarga nada e coletado ate ~19/09."),
    ("Continuo", "prospect-enrich-drain", "144 execucoes por dia",
     "Roda a cada 10 minutos enriquecendo leads para prospeccao, na MESMA conta Apify. E o maior "
     "consumidor: o enriquecimento de prospeccao come o orcamento da coleta de conteudo."),
]

for linha in DIAGNOSTICO:
    ws.append(list(linha))

body_font(ws, first_row=5)
for i in range(5, ws.max_row + 1):
    ws.cell(row=i, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[i].height = 46
    if "FALHOU" in str(ws.cell(row=i, column=3).value):
        ws.cell(row=i, column=3).fill = REV_FILL
widths(ws, {"A": 18, "B": 24, "C": 22, "D": 96})

consequencia = ws.max_row + 2
ws[f"A{consequencia}"] = "O QUE ISSO CUSTA NO CRUZAMENTO"
ws[f"A{consequencia}"].font = Font(name=FONT, size=10, bold=True, color="9E5714")
for col in range(1, 5):
    ws.cell(row=consequencia, column=col).fill = NOTE_FILL
ws[f"A{consequencia + 1}"] = (
    "Sem o post em content_posts nao existe linha em post_lead_magnets ligando o post ao "
    "formulario xX8rVJ - e esse vinculo e a evidencia mais forte do matcher (confianca 0,96). "
    "E por isso que 12 nomes ficaram em REVIEW em vez de confirmados. Dois pontos a mais: o CTA "
    "ROTINAS nao existe no mapa CTA_TO_TALLY_FORM (leadMagnets.ts), e o post nao pediu comentario, "
    "entao provavelmente vai entrar sem cta_keyword - o vinculo tera de ser inserido a mao em "
    "post_lead_magnets, que e a fonte da verdade em producao."
)
ws[f"A{consequencia + 1}"].font = Font(name=FONT, size=9, color=INK)
ws[f"A{consequencia + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{consequencia + 1}:D{consequencia + 1}")
ws.row_dimensions[consequencia + 1].height = 76

acao = consequencia + 3
ws[f"A{acao}"] = "PARA DESTRAVAR"
ws[f"A{acao}"].font = Font(name=FONT, size=10, bold=True, color="0F6E62")
ws[f"A{acao + 1}"] = (
    "1. Recarregar a Apify ou esperar o reset do ciclo - e decisao de custo, nao tem jeito tecnico.\n"
    "2. Enquanto isso, a prospeccao deste post nao depende da coleta: as inscricoes do Tally ja "
    "trazem nome, e-mail e telefone (aba Inscritos).\n"
    "3. Quando o post entrar, inserir o vinculo em post_lead_magnets e rodar de novo o cruzamento - "
    "boa parte dos REVIEW deve virar confirmado sozinha.\n"
    "4. Vale separar o orcamento de prospeccao do de conteudo, ou o enriquecimento vai continuar "
    "derrubando a coleta todo mes.\n"
    "5. Vale rodar o coletor de LinkedIn duas vezes ao dia: com uma rodada as 09:30 UTC, um post "
    "publicado ao meio-dia espera 21h para ser capturado."
)
ws[f"A{acao + 1}"].font = Font(name=FONT, size=9, color=INK)
ws[f"A{acao + 1}"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells(f"A{acao + 1}:D{acao + 1}")
ws.row_dimensions[acao + 1].height = 118

# ---------------------------------------------------------------- ordem das abas

wb.move_sheet("Resumo", offset=-wb.index(wb["Resumo"]))
for pos, nome in enumerate(["Resumo", "Inscritos", "Matches", "Comentarios", "Historico Tally", "ICP", "Enriquecer", "Diagnostico"]):
    atual = wb.index(wb[nome])
    wb.move_sheet(nome, offset=pos - atual)

wb.save(OUT)
print(f"Gravado: {OUT}")
print(f"Inscritos: {len(inscritos)} | matches: {len(matches)} | comentarios: {len(comentarios)} | "
      f"historico: {len(historico)} | enriquecer: {len(para_enriquecer)}")

# ---------------------------------------------------------------- conferencia
# Nao ha LibreOffice nesta maquina, entao as formulas nao podem ser recalculadas aqui.
# Em vez de confiar nelas no escuro, recalculo cada numero em Python puro sobre os
# mesmos dados e imprimo o par (esperado, criterio). Se os dois lados divergirem, o
# criterio da formula esta errado.

reais = [r for r in inscritos if r["classificacao"] == "Real"]
comentou = [r for r in reais if r["audiencia"] == "Comentou em post"]
tam = lambda r: r["tamanho_empresa"] if isinstance(r["tamanho_empresa"], (int, float)) else 0

conferencia = [
    ("Inscricoes no formulario", len(inscritos)),
    ("Pessoas reais", len(reais)),
    ("Com telefone (comeca com +)", sum(1 for r in reais if str(r["telefone"]).startswith("+"))),
    ("Com e-mail corporativo", sum(1 for r in reais if r["corporativo"] == "sim")),
    ("Comentaram em post nosso", len(comentou)),
    ("  MATCHED", sum(1 for r in reais if r["match_status"] == "MATCHED")),
    ("  REVIEW", sum(1 for r in reais if r["match_status"] == "REVIEW")),
    ("Baixaram outro lead magnet", sum(1 for r in reais if r["audiencia"] == "Baixou outro lead magnet")),
    ("Audiencia nova", sum(1 for r in reais if r["audiencia"] == "Novo")),
    ("Qualified", sum(1 for r in reais if r["icp_status"] == "qualified")),
    ("Disqualified", sum(1 for r in reais if r["icp_status"] == "disqualified")),
    ("Ja prospectados", sum(1 for r in reais if r["prospectado"] != "não")),
    ("Agendaram reuniao", sum(1 for r in reais if r["agendou_reuniao"] != "não")),
    ("--- cortes de ICP (so quem comentou) ---", ""),
    ("area = vendas", sum(1 for r in comentou if r["area"] == "vendas")),
    ("c-level", sum(1 for r in comentou if r["senioridade"] == "c-level")),
    ("diretoria + gerencia", sum(1 for r in comentou if r["senioridade"] in ("diretoria", "gerencia"))),
    ("operacional + coordenacao", sum(1 for r in comentou if r["senioridade"] in ("operacional", "coordenacao"))),
    ("senioridade desconhecida", sum(1 for r in comentou if r["senioridade"] == "desconhecido")),
    ("empresa > 1000", sum(1 for r in comentou if tam(r) > 1000)),
    ("empresa 11..1000", sum(1 for r in comentou if 11 <= tam(r) <= 1000)),
    ("empresa 1..10", sum(1 for r in comentou if 0 < tam(r) <= 10)),
    ("sem empresa", sum(1 for r in comentou if tam(r) == 0)),
    ("2+ comentarios", sum(1 for r in comentou if r["comentarios"] >= 2)),
    ("tambem baixou outro", sum(1 for r in comentou if r["outros_lead_magnets"] > 0)),
]

print("\nConferencia (valor esperado de cada formula, calculado em Python):")
for label, valor in conferencia:
    print(f"  {label:44} {valor}")

# Sanidade estrutural (nao fixa numeros: o Tally sincroniza de hora em hora e a base
# cresce). So garante coerencia interna do que foi montado.
matched = sum(1 for r in reais if r["match_status"] == "MATCHED")
review = sum(1 for r in reais if r["match_status"] == "REVIEW")
assert len(reais) > 0 and len(reais) <= len(inscritos)
assert len(comentou) == matched + review + sum(1 for r in reais
    if r["audiencia"] == "Comentou em post" and r["match_status"] in ("MATCHED_NO_PHONE", "NOT_FOUND"))
assert matched + review <= len(comentou)
print(f"\nSanidade ok: {len(reais)} reais, {len(comentou)} comentaram, "
      f"{matched} MATCHED, {review} REVIEW.")
